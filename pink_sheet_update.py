"""
pink_sheet_update.py
====================
Downloads the World Bank Commodity Price Data ("Pink Sheet") monthly
Excel file, parses all commodity series, and writes pink_sheet.json
to the same directory for use by index.html (madadim.net).

The script auto-discovers the current XLS URL by scraping the WB
commodity markets page, so it stays up to date even when the World Bank
rotates document IDs.

Run any time you want fresh data:
    python pink_sheet_update.py

Override the output path:
    python pink_sheet_update.py --out /path/to/pink_sheet.json

Force a specific XLS URL (skips auto-discovery):
    python pink_sheet_update.py --url https://thedocs.worldbank.org/.../CMO-Historical-Data-Monthly.xlsx

Requirements:
    pip install requests openpyxl beautifulsoup4
"""

import argparse
import io
import json
import os
import re
import sys
from datetime import datetime, date

import requests
import openpyxl
from bs4 import BeautifulSoup

# ── Config ────────────────────────────────────────────────────────────────────

COMMODITY_PAGE_URL = "https://www.worldbank.org/en/research/commodity-markets"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}

CATEGORIES = [
    ("crude",       "Energy"),
    ("coal",        "Energy"),
    ("natural gas", "Energy"),
    ("lng",         "Energy"),
    ("petrol",      "Energy"),
    ("copper",      "Metals"),
    ("aluminum",    "Metals"),
    ("aluminium",   "Metals"),
    ("iron ore",    "Metals"),
    ("tin",         "Metals"),
    ("nickel",      "Metals"),
    ("zinc",        "Metals"),
    ("lead",        "Metals"),
    ("gold",        "Metals"),
    ("silver",      "Metals"),
    ("platinum",    "Metals"),
    ("wheat",       "Agriculture"),
    ("maize",       "Agriculture"),
    ("corn",        "Agriculture"),
    ("rice",        "Agriculture"),
    ("sorghum",     "Agriculture"),
    ("barley",      "Agriculture"),
    ("palm",        "Agriculture"),
    ("soybean",     "Agriculture"),
    ("coconut",     "Agriculture"),
    ("groundnut",   "Agriculture"),
    ("sunflower",   "Agriculture"),
    ("coffee",      "Agriculture"),
    ("cocoa",       "Agriculture"),
    ("tea",         "Agriculture"),
    ("banana",      "Agriculture"),
    ("orange",      "Agriculture"),
    ("sugar",       "Agriculture"),
    ("cotton",      "Agriculture"),
    ("rubber",      "Agriculture"),
    ("tobacco",     "Agriculture"),
    ("timber",      "Agriculture"),
    ("logs",        "Agriculture"),
    ("sawnwood",    "Agriculture"),
    ("dap",         "Fertilizers"),
    ("urea",        "Fertilizers"),
    ("potassium",   "Fertilizers"),
    ("tsp",         "Fertilizers"),
    ("phosphate",   "Fertilizers"),
]


def categorize(name: str) -> str:
    lower = name.lower()
    for keyword, cat in CATEGORIES:
        if keyword in lower:
            return cat
    return "Other"


def make_code(name: str) -> str:
    clean = re.sub(r"[^a-zA-Z0-9]+", "_", name.strip()).upper().strip("_")
    return f"WB_{clean}"


# ── URL discovery ─────────────────────────────────────────────────────────────

def discover_xlsx_url(session: requests.Session) -> str:
    """
    Scrape the WB commodity markets page to find the current monthly
    prices XLS URL. Targets the <a> whose text contains both 'monthly'
    and 'price' and whose href ends in .xlsx — matching the 'Monthly prices'
    cell in the data table.
    """
    print(f"  Fetching index page: {COMMODITY_PAGE_URL}")
    r = session.get(COMMODITY_PAGE_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")
    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True).lower()
        href = a["href"]
        if "monthly" in text and "price" in text and href.endswith(".xlsx"):
            # Ensure absolute URL
            if href.startswith("http"):
                return href
            return "https://thedocs.worldbank.org" + href

    raise ValueError(
        "Could not find a 'Monthly prices' .xlsx link on the WB commodity markets page.\n"
        "The page structure may have changed. Use --url to supply the link manually."
    )


# ── Download ──────────────────────────────────────────────────────────────────

def download(url: str, session: requests.Session) -> bytes:
    print(f"  GET {url}")
    r = session.get(url, headers=HEADERS, timeout=90, stream=True)
    r.raise_for_status()
    total = int(r.headers.get("content-length", 0))
    buf = []
    done = 0
    for chunk in r.iter_content(65536):
        buf.append(chunk)
        done += len(chunk)
        if total:
            print(
                f"\r  {done // 1024} KB / {total // 1024} KB  ({done / total * 100:.0f}%)",
                end="", flush=True,
            )
    print()
    return b"".join(buf)


# ── Date parsing ──────────────────────────────────────────────────────────────

def parse_date_cell(val) -> str | None:
    """
    Parse date cells that appear in the Pink Sheet.
    Handles: datetime objects, 'YYYYMxx', 'YYYY Mxx', Excel serial floats,
    'Jan-60', 'Jan-1960', 'January 1960', and ISO strings.
    Returns 'YYYY-MM' or None.
    """
    if val is None:
        return None

    # openpyxl may return a proper datetime for date-formatted cells
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m")

    # Excel serial number (integer or float, plausible date range)
    if isinstance(val, (int, float)) and 15000 < val < 80000:
        from datetime import timedelta
        dt = datetime(1899, 12, 30) + timedelta(days=int(val))
        return dt.strftime("%Y-%m")

    s = str(val).strip()

    # "1960M01" or "1960 M01" — the actual Pink Sheet format
    m = re.match(r"(\d{4})\s*[Mm](\d{1,2})$", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"

    # Various other formats the WB has used historically
    for fmt in ("%b-%y", "%b-%Y", "%B %Y", "%B-%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            # Two-digit year: 60-99 → 1960-1999, 00-59 → 2000-2059
            if dt.year < 100:
                dt = dt.replace(year=dt.year + (1900 if dt.year >= 60 else 2000))
            return dt.strftime("%Y-%m")
        except ValueError:
            continue

    return None


# ── Sheet parsing ─────────────────────────────────────────────────────────────

def find_monthly_sheet(wb: openpyxl.Workbook) -> str:
    for name in wb.sheetnames:
        if "monthly" in name.lower() and "price" in name.lower():
            return name
    for name in wb.sheetnames:
        if "monthly" in name.lower():
            return name
    return wb.sheetnames[0]


def parse_pink_sheet(xlsx_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheet_name = find_monthly_sheet(wb)
    ws = wb[sheet_name]
    print(f"  Sheet: '{sheet_name}'")

    rows = list(ws.iter_rows(values_only=True))

    # ── Locate the unit row and data start ────────────────────────────────────
    # The unit row is the last row (before data) where any cell in cols B–H
    # contains "$". The header row is immediately above it. Data starts on the
    # first row after the unit row whose col A parses as a date.

    unit_row_idx = None
    for i, row in enumerate(rows):
        for cell in row[1:8]:
            if cell is not None and "$" in str(cell):
                unit_row_idx = i
                break

    if unit_row_idx is None:
        raise ValueError("Cannot find unit row (no cell containing '$' in cols B-H).")

    data_start_idx = None
    for i in range(unit_row_idx + 1, len(rows)):
        if parse_date_cell(rows[i][0]) is not None:
            data_start_idx = i
            break

    if data_start_idx is None:
        raise ValueError("Cannot find first data row with a parseable date in col A.")

    header_row_idx = unit_row_idx - 1
    print(
        f"  Header row: {header_row_idx + 1}  |  "
        f"Unit row: {unit_row_idx + 1}  |  "
        f"Data from row: {data_start_idx + 1}"
    )

    headers = rows[header_row_idx]
    units   = rows[unit_row_idx]

    # ── Build series index from header row ────────────────────────────────────
    series_list = []
    col_map = {}  # col_index → position in series_list

    for col_i, (h, u) in enumerate(zip(headers, units)):
        if col_i == 0 or h is None:
            continue
        name = str(h).strip()
        if not name or name.lower() == "nan":
            continue
        unit = str(u).strip() if u is not None else ""
        series_list.append({
            "code":     make_code(name),
            "name":     name,
            "unit":     unit,
            "category": categorize(name),
            "data":     [],
        })
        col_map[col_i] = len(series_list) - 1

    print(f"  Commodity columns found: {len(series_list)}")

    # ── Parse data rows ───────────────────────────────────────────────────────
    for row in rows[data_start_idx:]:
        date_str = parse_date_cell(row[0])
        if date_str is None:
            continue  # skip metadata/footer rows (e.g. "Updated on June 02, 2026")
        for col_i, series_idx in col_map.items():
            if col_i >= len(row):
                continue
            val = row[col_i]
            if val is None:
                continue
            try:
                fval = float(val)
                if fval > 0:
                    series_list[series_idx]["data"].append({
                        "date":  date_str,
                        "value": round(fval, 4),
                    })
            except (TypeError, ValueError):
                continue

    # Sort and drop empty series
    for s in series_list:
        s["data"].sort(key=lambda x: x["date"])
    series_list = [s for s in series_list if s["data"]]

    print(f"  Series with data: {len(series_list)}")
    return series_list


# ── Year-on-year % change ─────────────────────────────────────────────────────

def add_pct_year(series_list: list[dict]) -> None:
    for s in series_list:
        by_date = {r["date"]: r["value"] for r in s["data"]}
        for row in s["data"]:
            yr, mo = row["date"].split("-")
            prev_date = f"{int(yr) - 1}-{mo}"
            prev_val  = by_date.get(prev_date)
            if prev_val and prev_val > 0:
                row["pct_year"] = round((row["value"] / prev_val - 1) * 100, 2)
            else:
                row["pct_year"] = None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="World Bank Pink Sheet → pink_sheet.json"
    )
    parser.add_argument(
        "--url", default=None,
        help=(
            "Direct XLS URL (skips auto-discovery). Use this if the WB page "
            "structure breaks discovery: python pink_sheet_update.py --url <url>"
        ),
    )
    parser.add_argument(
        "--out", default=None,
        help="Output path (default: pink_sheet.json next to this script)",
    )
    args = parser.parse_args()

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = args.out or os.path.join(base_dir, "pink_sheet.json")

    print("\nWorld Bank Pink Sheet Updater")
    print(f"  Output : {out_path}\n")

    session = requests.Session()

    # ── Step 1: resolve XLS URL ───────────────────────────────────────────────
    if args.url:
        xlsx_url = args.url
        print(f"Step 1 — Using supplied URL:\n  {xlsx_url}\n")
    else:
        print("Step 1 — Discovering current XLS URL from WB commodity markets page...")
        try:
            xlsx_url = discover_xlsx_url(session)
        except Exception as e:
            print(f"\n  [ERROR] {e}", file=sys.stderr)
            print(
                "  Tip: supply the URL manually with --url and report the issue.",
                file=sys.stderr,
            )
            sys.exit(1)
        print(f"  URL: {xlsx_url}\n")

    # ── Step 2: download ──────────────────────────────────────────────────────
    print("Step 2 — Downloading Excel file...")
    try:
        xlsx_bytes = download(xlsx_url, session)
    except requests.HTTPError as e:
        print(
            f"\n  [ERROR] HTTP {e.response.status_code}.\n"
            "  The discovered URL may be stale. Try running with --url and the "
            "latest link from:\n"
            f"  {COMMODITY_PAGE_URL}",
            file=sys.stderr,
        )
        sys.exit(1)
    except Exception as e:
        print(f"\n  [ERROR] {e}", file=sys.stderr)
        sys.exit(1)
    print(f"  Size: {len(xlsx_bytes) // 1024} KB")

    # ── Step 3: parse ─────────────────────────────────────────────────────────
    print("\nStep 3 — Parsing monthly prices sheet...")
    try:
        series_list = parse_pink_sheet(xlsx_bytes)
    except Exception as e:
        print(f"  [ERROR] {e}", file=sys.stderr)
        sys.exit(1)

    # ── Step 4: year-on-year changes ──────────────────────────────────────────
    print("\nStep 4 — Computing year-on-year % changes...")
    add_pct_year(series_list)

    # ── Step 5: write JSON ────────────────────────────────────────────────────
    last_date = max(
        (row["date"] for s in series_list for row in s["data"]),
        default="unknown",
    )
    output = {
        "updated":    datetime.now().strftime("%Y-%m-%d"),
        "last_date":  last_date,
        "source":     "World Bank Commodity Price Data (Pink Sheet)",
        "source_url": xlsx_url,
        "series":     series_list,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, separators=(",", ":"))

    total_pts = sum(len(s["data"]) for s in series_list)
    size_kb   = os.path.getsize(out_path) // 1024

    print(f"\n  Saved  : {out_path}")
    print(f"  Series : {len(series_list)}")
    print(f"  Points : {total_pts:,}")
    print(f"  Size   : {size_kb} KB")
    print(f"  Through: {last_date}")
    print(f"  Run on : {output['updated']}\n")


if __name__ == "__main__":
    main()
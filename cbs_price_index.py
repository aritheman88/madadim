"""
cbs_price_index.py
==================
Fetches monthly price index data from the Israeli CBS API and generates:
  - An interactive self-contained HTML file
  - An Excel workbook (.xlsx) with the raw data

Indices fetched:
    120010 — מדד המחירים לצרכן (CPI General)
    120050 — מזון ללא פירות וירקות (Food excl. fruit & veg)

The CBS API caps responses at ~100 rows and periodically rebases the index
(values reset to ~100 each new base period). This script:
  - Fetches in 5-year chunks to get the full history
  - Detects rebase boundaries and chains values into one continuous series

Usage:
    python cbs_price_index.py
    python cbs_price_index.py --from 2010-01 --to 2025-12
    python cbs_price_index.py --out my_report

Requirements:
    pip install requests python-dateutil openpyxl
    Must be run from an Israeli IP address.
"""

import argparse
import json
import os
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime
from dateutil.relativedelta import relativedelta

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# ── Config ────────────────────────────────────────────────────────────────────

API_BASE = "https://api.cbs.gov.il/index/data/price"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/xml, text/xml, */*",
}

INDICES = {
    "cpi": {
        "id":    "120010",
        "label": "מדד המחירים לצרכן",
        "color": "#f59e0b",
    },
    "food": {
        "id":    "120050",
        "label": "מזון ללא פירות וירקות",
        "color": "#38bdf8",
    },
}

CHUNK_MONTHS = 60        # fetch 5 years at a time to stay under API row cap

# ── Fetch ─────────────────────────────────────────────────────────────────────

def ym_to_date(ym):
    return datetime.strptime(ym, "%Y-%m")

def date_to_ym(dt):
    return dt.strftime("%Y-%m")

def cbs_period(ym):
    """Convert YYYY-MM → MM-YYYY (CBS API format)."""
    y, m = ym.split("-")
    return f"{m}-{y}"

def fetch_chunk(index_id, from_ym, to_ym):
    params = {
        "id":          index_id,
        "startPeriod": cbs_period(from_ym),
        "endPeriod":   cbs_period(to_ym),
        "format":      "xml",
        "lang":        "he",
        "baseType":    "1",
    }
    try:
        r = requests.get(API_BASE, params=params, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"    [ERROR] {e}", file=sys.stderr)
        return []

    text = r.text.strip()
    if not text.startswith("<"):
        print(f"    [ERROR] Non-XML response: {text[:150]}", file=sys.stderr)
        return []

    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        print(f"    [ERROR] XML parse error: {e}", file=sys.stderr)
        return []

    rows = []
    for dm in root.iter("DateMonth"):
        def txt(tag):
            el = dm.find(tag)
            return (el.text or "").strip() if el is not None else ""

        year  = txt("year")
        month = txt("month").zfill(2)
        base  = dm.find("currBase")
        value_str = ""
        if base is not None:
            ve = base.find("value")
            value_str = (ve.text or "").strip() if ve is not None else ""
        pct_str = txt("percentYear")

        if not (year and month and value_str):
            continue
        try:
            value = float(value_str)
        except ValueError:
            continue

        pct_year = None
        try:
            pct_year = float(pct_str)
        except ValueError:
            pass

        rows.append({"date": f"{year}-{month}", "value": value, "pct_year": pct_year})

    rows.sort(key=lambda r: r["date"])
    return rows


def fetch_all_chunks(index_id, from_ym, to_ym):
    """
    Fetch in CHUNK_MONTHS-sized windows and concatenate.
    Deduplicates on date. Returns raw values as-is from the API.
    """
    all_rows = []
    cursor   = ym_to_date(from_ym)
    end      = ym_to_date(to_ym)
    seen     = set()

    while cursor <= end:
        chunk_end = min(cursor + relativedelta(months=CHUNK_MONTHS - 1), end)
        from_s = date_to_ym(cursor)
        to_s   = date_to_ym(chunk_end)
        print(f"    chunk {from_s} → {to_s} ...", end=" ", flush=True)

        rows = fetch_chunk(index_id, from_s, to_s)
        new  = [r for r in rows if r["date"] not in seen]
        seen.update(r["date"] for r in new)
        all_rows.extend(new)
        print(f"{len(new)} rows")

        cursor = chunk_end + relativedelta(months=1)
        time.sleep(0.4)

    all_rows.sort(key=lambda r: r["date"])
    return all_rows


def rescale_to_base2010(rows):
    """
    1. Chain across January rebase boundaries: the CBS API resets its base
       every few years, always in January. We detect these as a January value
       that is lower than the previous December by more than 1 point, and scale
       all subsequent rows upward to eliminate the discontinuity.
    2. Rescale so the annual average of 2010 = 100 (CBS convention).
    """
    if not rows:
        return rows

    rows = [dict(r) for r in rows]

    # Step 1: chain January rebase drops
    for i in range(1, len(rows)):
        if rows[i]["date"].endswith("-01"):
            prev_val = rows[i-1]["value"]
            curr_val = rows[i]["value"]
            drop = prev_val - curr_val
            if drop > 1.0:   # more than 1 index point drop at January = rebase
                scale = prev_val / curr_val
                print(f"    rebase at {rows[i]['date']}: {prev_val:.3f} → {curr_val:.3f}  (×{scale:.4f})")
                for j in range(i, len(rows)):
                    rows[j]["value"] = rows[j]["value"] * scale

    # Step 2: rescale to annual average 2010 = 100
    vals_2010 = [r["value"] for r in rows if r["date"].startswith("2010")]
    if not vals_2010:
        print("    [WARN] No 2010 data — skipping rebase")
        return rows
    avg_2010 = sum(vals_2010) / len(vals_2010)
    print(f"    2010 average = {avg_2010:.4f} → rescaling to base 2010=100")
    return [dict(r, value=round(r["value"] / avg_2010 * 100, 3)) for r in rows]


# ── XLSX output ───────────────────────────────────────────────────────────────

def generate_xlsx(all_data, from_ym, to_ym, out_path):
    cpi_rows  = all_data["cpi"]
    food_rows = all_data["food"]

    all_dates    = sorted(set(r["date"] for r in cpi_rows) | set(r["date"] for r in food_rows))
    cpi_by_date  = {r["date"]: r for r in cpi_rows}
    food_by_date = {r["date"]: r for r in food_rows}
    cpi_base     = cpi_rows[0]["value"]  if cpi_rows  else 1
    food_base    = food_rows[0]["value"] if food_rows else 1

    wb = openpyxl.Workbook()

    # ── Sheet 1: Raw chained values ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "נתונים גולמיים"

    hdr_fill  = PatternFill("solid", fgColor="1F2937")
    cpi_fill  = PatternFill("solid", fgColor="92400E")
    food_fill = PatternFill("solid", fgColor="0C4A6E")
    bold_w    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    normal    = Font(name="Arial", size=10)

    headers = [
        "תאריך",
        "CPI — ערך (שרשרת)", "CPI — % שנתי",
        "מזון — ערך (שרשרת)", "מזון — % שנתי",
    ]
    fills = [hdr_fill, cpi_fill, cpi_fill, food_fill, food_fill]
    for col, (h, fill) in enumerate(zip(headers, fills), 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = bold_w
        c.fill = fill
        c.alignment = Alignment(horizontal="center")

    for row_i, date in enumerate(all_dates, 2):
        cpi  = cpi_by_date.get(date)
        food = food_by_date.get(date)
        row_vals = [
            date,
            cpi["value"]    if cpi  else None,
            cpi["pct_year"] if cpi  else None,
            food["value"]   if food else None,
            food["pct_year"]if food else None,
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws1.cell(row=row_i, column=col, value=val)
            c.font = normal
            c.alignment = Alignment(horizontal="center" if col == 1 else "right")
            if isinstance(val, float):
                c.number_format = "0.00"

    ws1.column_dimensions["A"].width = 12
    for col in ["B", "C", "D", "E"]:
        ws1.column_dimensions[col].width = 20
    ws1.freeze_panes = "A2"

    # ── Sheet 2: % change from base period ───────────────────────────────────
    ws2 = wb.create_sheet("% מבסיס")

    for col, (h, fill) in enumerate(zip(
        ["תאריך", f"CPI % מ-{from_ym}", f"מזון % מ-{from_ym}"],
        [hdr_fill, cpi_fill, food_fill]
    ), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = bold_w
        c.fill = fill
        c.alignment = Alignment(horizontal="center")

    for row_i, date in enumerate(all_dates, 2):
        cpi  = cpi_by_date.get(date)
        food = food_by_date.get(date)
        cpi_chg  = round((cpi["value"]  / cpi_base  - 1) * 100, 2) if cpi  else None
        food_chg = round((food["value"] / food_base - 1) * 100, 2) if food else None
        for col, val in enumerate([date, cpi_chg, food_chg], 1):
            c = ws2.cell(row=row_i, column=col, value=val)
            c.font = normal
            c.alignment = Alignment(horizontal="center" if col == 1 else "right")
            if isinstance(val, float):
                c.number_format = "0.00"

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 20
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"  → Saved: {out_path}")


# ── HTML output ───────────────────────────────────────────────────────────────

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CBS Price Index Explorer</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-adapter-date-fns@3/dist/chartjs-adapter-date-fns.bundle.min.js"></script>
<link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans+Hebrew:wght@300;400;600&display=swap" rel="stylesheet">
<style>
  :root { --bg:#0d0f14; --surface:#161922; --border:#252a35; --text:#e2e8f0; --muted:#6b7585; --accent:#4ade80; }
  * { box-sizing:border-box; margin:0; padding:0; }
  body { background:var(--bg); color:var(--text); font-family:'IBM Plex Sans Hebrew',sans-serif; min-height:100vh; display:flex; flex-direction:column; }
  header { border-bottom:1px solid var(--border); padding:18px 32px; display:flex; align-items:baseline; gap:16px; flex-wrap:wrap; }
  header h1 { font-family:'IBM Plex Mono',monospace; font-size:1.05rem; font-weight:600; color:var(--accent); letter-spacing:.05em; }
  header p  { font-size:.78rem; color:var(--muted); font-family:'IBM Plex Mono',monospace; }
  .controls { padding:16px 32px; display:flex; gap:20px; align-items:flex-end; flex-wrap:wrap; border-bottom:1px solid var(--border); }
  .control-group { display:flex; flex-direction:column; gap:6px; }
  .control-group label { font-size:.68rem; color:var(--muted); font-family:'IBM Plex Mono',monospace; text-transform:uppercase; letter-spacing:.08em; }
  .range-inputs { display:flex; align-items:center; gap:8px; }
  input[type="month"] { background:var(--surface); border:1px solid var(--border); color:var(--text); padding:7px 12px; font-family:'IBM Plex Mono',monospace; font-size:.83rem; border-radius:4px; outline:none; }
  .range-sep { color:var(--muted); font-family:'IBM Plex Mono',monospace; font-size:.8rem; }
  .toggles { display:flex; gap:10px; }
  .toggle-btn { display:flex; align-items:center; gap:8px; padding:7px 14px; border-radius:4px; border:1px solid var(--border); background:var(--surface); color:var(--muted); cursor:pointer; font-family:'IBM Plex Mono',monospace; font-size:.78rem; transition:all .2s; user-select:none; }
  .toggle-btn .swatch { width:9px; height:9px; border-radius:2px; opacity:.3; transition:opacity .2s; }
  .toggle-btn.active { color:var(--col); border-color:var(--col); }
  .toggle-btn.active .swatch { opacity:1; }
  .mode-group { display:flex; gap:4px; }
  .mode-btn { padding:7px 13px; border-radius:4px; border:1px solid var(--border); background:var(--surface); color:var(--muted); cursor:pointer; font-family:'IBM Plex Mono',monospace; font-size:.75rem; transition:all .2s; }
  .mode-btn.active { background:var(--accent); color:var(--bg); border-color:var(--accent); font-weight:600; }
  .chart-wrap { flex:1; padding:24px 32px; height:460px; }
  canvas { width:100% !important; height:100% !important; }
  .stats-row { display:flex; gap:1px; background:var(--border); border-top:1px solid var(--border); }
  .stat-card { flex:1; background:var(--surface); padding:14px 22px; display:flex; flex-direction:column; gap:3px; }
  .stat-label { font-size:.66rem; color:var(--muted); font-family:'IBM Plex Mono',monospace; text-transform:uppercase; letter-spacing:.06em; }
  .stat-value { font-size:1.35rem; font-family:'IBM Plex Mono',monospace; font-weight:600; }
  .stat-sub { font-size:.7rem; color:var(--muted); font-family:'IBM Plex Mono',monospace; }
  .footer { text-align:center; font-size:.68rem; color:var(--muted); font-family:'IBM Plex Mono',monospace; padding:10px; border-top:1px solid var(--border); }
</style>
</head>
<body>
<header>
  <h1>CBS // Price Index Explorer</h1>
  <p>הלשכה המרכזית לסטטיסטיקה · Generated: __GENERATED__</p>
</header>
<div class="controls">
  <div class="control-group">
    <label>טווח תאריכים</label>
    <div class="range-inputs">
      <input type="month" id="fromDate" value="__FROM__">
      <span class="range-sep">→</span>
      <input type="month" id="toDate" value="__TO__">
    </div>
  </div>
  <div class="control-group">
    <label>מדדים</label>
    <div class="toggles" id="toggles"></div>
  </div>
  <div class="control-group">
    <label>תצוגה</label>
    <div class="mode-group">
      <button class="mode-btn"        onclick="setMode('index',this)">ערך (שרשרת)</button>
      <button class="mode-btn"        onclick="setMode('pct',this)">% שנתי</button>
      <button class="mode-btn active" onclick="setMode('base',this)">% מבסיס</button>
    </div>
  </div>
</div>
<div class="chart-wrap"><canvas id="myChart"></canvas></div>
<div class="stats-row" id="statsRow"></div>
<div class="footer">__TOTAL_POINTS__ data points · values chained across rebase boundaries</div>

<script>
const ALL_DATA = __ALL_DATA_JSON__;
const INDICES  = __INDICES_JSON__;

let mode    = "base";
let visible = Object.fromEntries(Object.keys(INDICES).map(k => [k, true]));
let chart   = null;

// toggles
const togglesEl = document.getElementById("toggles");
Object.entries(INDICES).forEach(([key, meta]) => {
  const btn = document.createElement("button");
  btn.className = "toggle-btn active";
  btn.style.setProperty("--col", meta.color);
  btn.innerHTML = `<span class="swatch" style="background:${meta.color}"></span>${meta.label}`;
  btn.onclick = () => { visible[key] = !visible[key]; btn.classList.toggle("active", visible[key]); updateChart(); };
  togglesEl.appendChild(btn);
});

// stats
const statsRow = document.getElementById("statsRow");
Object.entries(INDICES).forEach(([key, meta]) => {
  const rows = ALL_DATA[key];
  const last = rows[rows.length - 1], first = rows[0];
  const gain = ((last.value / first.value - 1) * 100).toFixed(1);
  statsRow.innerHTML += `
    <div class="stat-card">
      <span class="stat-label" style="color:${meta.color}">${meta.label} · ערך אחרון</span>
      <span class="stat-value" style="color:${meta.color}">${last.value.toFixed(1)}</span>
      <span class="stat-sub">${last.date}</span>
    </div>
    <div class="stat-card">
      <span class="stat-label" style="color:${meta.color}">${meta.label} · % שנתי</span>
      <span class="stat-value" style="color:${meta.color}">${last.pct_year != null ? (last.pct_year > 0 ? "+" : "") + last.pct_year.toFixed(1) + "%" : "—"}</span>
      <span class="stat-sub">year-on-year</span>
    </div>
    <div class="stat-card">
      <span class="stat-label" style="color:${meta.color}">${meta.label} · עלייה כוללת</span>
      <span class="stat-value" style="color:${meta.color}">+${gain}%</span>
      <span class="stat-sub">${first.date} → ${last.date}</span>
    </div>`;
});

function filtered(key) {
  const from = document.getElementById("fromDate").value;
  const to   = document.getElementById("toDate").value;
  return ALL_DATA[key].filter(r => r.date >= from && r.date <= to);
}

function buildSeries(key) {
  const rows = filtered(key);
  if (!rows.length) return [];
  // x must be a Date object for the time scale
  if (mode === "index") return rows.map(r => ({ x: new Date(r.date + "-01"), y: r.value }));
  if (mode === "pct")   return rows.map(r => ({ x: new Date(r.date + "-01"), y: r.pct_year }));
  if (mode === "base") {
    const base = rows[0].value;
    return rows.map(r => ({ x: new Date(r.date + "-01"), y: +((r.value / base - 1) * 100).toFixed(2) }));
  }
}

function updateChart() {
  const datasets = Object.entries(INDICES)
    .filter(([k]) => visible[k])
    .map(([k, meta]) => ({
      label: meta.label, data: buildSeries(k),
      borderColor: meta.color, backgroundColor: meta.color + "18",
      borderWidth: 1.8, pointRadius: 0, pointHoverRadius: 4,
      tension: 0.2, fill: false,
    }));

  const yLabel = mode === "index" ? "ערך מדד (שרשרת)" : mode === "pct" ? "% שינוי שנתי" : "% מבסיס";

  if (chart) {
    chart.data.datasets = datasets;
    chart.options.scales.y.title.text = yLabel;
    chart.update();
    return;
  }

  chart = new Chart(document.getElementById("myChart"), {
    type: "line",
    data: { datasets },
    options: {
      responsive: true, maintainAspectRatio: false, animation: { duration: 300 },
      interaction: { mode: "index", intersect: false },
      plugins: {
        legend: { display: true, labels: { color: "#e2e8f0", font: { family: "'IBM Plex Mono', monospace", size: 11 }, boxWidth: 12 } },
        tooltip: {
          backgroundColor: "#161922", borderColor: "#252a35", borderWidth: 1,
          titleColor: "#6b7585", bodyColor: "#e2e8f0",
          titleFont: { family: "'IBM Plex Mono', monospace", size: 11 },
          bodyFont:  { family: "'IBM Plex Mono', monospace", size: 12 },
          padding: 12,
          callbacks: { label: ctx => { const v = ctx.parsed.y; return v == null ? "N/A" : `${ctx.dataset.label}: ${v.toFixed(2)}${mode === "index" ? "" : "%"}`; } }
        }
      },
      scales: {
        x: {
          type: "time",
          time: { unit: "year", tooltipFormat: "yyyy-MM", displayFormats: { year: "yyyy", month: "MMM yyyy" } },
          ticks: { color: "#6b7585", font: { family: "'IBM Plex Mono', monospace", size: 10 }, maxRotation: 0 },
          grid: { color: "#252a35" },
        },
        y: {
          ticks: { color: "#6b7585", font: { family: "'IBM Plex Mono', monospace", size: 10 } },
          grid:  { color: "#252a35" },
          title: { display: true, text: yLabel, color: "#6b7585", font: { family: "'IBM Plex Mono', monospace", size: 10 } },
          afterDataLimits(axis) {
            const range = axis.max - axis.min;
            axis.min = axis.min - range * 0.05;
            axis.max = axis.max + range * 0.05;
          }
        }
      }
    }
  });
}

function setMode(m, btn) {
  mode = m;
  document.querySelectorAll(".mode-btn").forEach(b => b.classList.remove("active"));
  btn.classList.add("active");
  updateChart();
}

document.getElementById("fromDate").addEventListener("change", updateChart);
document.getElementById("toDate").addEventListener("change",   updateChart);
updateChart();
</script>
</body>
</html>
"""


def generate_html(all_data, from_ym, to_ym, out_path):
    total_points = sum(len(v) for v in all_data.values())
    html = HTML_TEMPLATE
    html = html.replace("__GENERATED__",     datetime.now().strftime("%Y-%m-%d %H:%M"))
    html = html.replace("__FROM__",          from_ym)
    html = html.replace("__TO__",            to_ym)
    html = html.replace("__TOTAL_POINTS__",  str(total_points))
    html = html.replace("__ALL_DATA_JSON__", json.dumps(all_data, ensure_ascii=False))
    html = html.replace("__INDICES_JSON__",  json.dumps(INDICES,  ensure_ascii=False))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  → Saved: {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CBS Price Index → HTML + XLSX")
    parser.add_argument("--from", dest="from_ym", default="2010-01",
                        help="Start month YYYY-MM (default: 2010-01)")
    parser.add_argument("--to",   dest="to_ym",
                        default=datetime.now().strftime("%Y-%m"),
                        help="End month YYYY-MM (default: current month)")
    parser.add_argument("--out",  default="cbs_price_index",
                        help="Output filename base, no extension (default: cbs_price_index)")
    args = parser.parse_args()

    base_dir  = os.path.dirname(os.path.abspath(__file__))
    html_path = os.path.join(base_dir, args.out + ".html")
    xlsx_path = os.path.join(base_dir, args.out + ".xlsx")

    print(f"\nCBS Price Index Fetcher")
    print(f"  Period : {args.from_ym} → {args.to_ym}")
    print(f"  Output : {args.out}.html + {args.out}.xlsx\n")

    all_data = {}
    for key, meta in INDICES.items():
        print(f"[{meta['id']}] {meta['label']}")
        rows = fetch_all_chunks(meta["id"], args.from_ym, args.to_ym)
        if not rows:
            print(f"  [WARN] No data — check Israeli IP")
            all_data[key] = []
            continue
        print(f"  {len(rows)} months fetched — rescaling to base 2010=100...")
        rows = rescale_to_base2010(rows)
        print(f"  Done. {rows[0]['date']} → {rows[-1]['date']}  "
              f"({rows[0]['value']:.1f} → {rows[-1]['value']:.1f}, "
              f"+{(rows[-1]['value']/rows[0]['value']-1)*100:.1f}%)")
        all_data[key] = rows

    print("\nGenerating outputs...")
    generate_html(all_data, args.from_ym, args.to_ym, html_path)
    generate_xlsx(all_data, args.from_ym, args.to_ym, xlsx_path)
    print("Done.\n")


if __name__ == "__main__":
    main()